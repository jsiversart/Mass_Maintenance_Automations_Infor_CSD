# nla_sup/pt2.py

import sys
from pathlib import Path

# --- Ensure repo root is in Python path ---
repo_root = Path(__file__).resolve().parent
while not (repo_root / "core").exists() and repo_root.parent != repo_root:
    repo_root = repo_root.parent

sys.path.insert(0, str(repo_root))


import os
import glob
import pandas as pd
import sqlite3
from pathlib import Path
from core.notifier import send_email
from core.config import PATHS, EMAILS
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows





# === HELPERS ===
def get_newest_file(prefix, folder=PATHS["sups_nlas_data"]):
    files = glob.glob(os.path.join(folder, f"{prefix}*"))
    return max(files, key=os.path.getctime) if files else None


def load_lookup(file_path):
    """Load old/new mapping from NLAs and SUPs Excel."""
    df = pd.read_excel(file_path, usecols=[0, 1], names=["old", "new"], dtype=str)
    df = df.dropna().astype(str)
    df["old"] = df["old"].str.strip()
    df["new"] = df["new"].str.strip()
    return df



def append_to_review_log(lookup_df, review_path):
    """
    Append processed mappings to a review log while retaining existing Excel formatting.
    """

    # --- Reload full lookup to access SKIP column ---
    full_lookup = pd.read_excel(
        PATHS["sups_nlas_data"] / "NLAs and SUPs to update via SAAMM.xlsx",
        usecols=[0, 1, 3, 4],
        names=["OLD", "NEW", "NOTES", "SKIP"],
        dtype=str
    )
    full_lookup = full_lookup.apply(lambda x: x.str.strip().str.upper() if x.dtype == "object" else x)
    lookup_df.columns = lookup_df.columns.str.upper()

    # --- Merge to attach skip info ---
    review_df = pd.merge(lookup_df, full_lookup, on=["OLD", "NEW"], how="left")
    review_df = review_df[review_df["SKIP"] != "Y"].copy()
    review_df.drop(columns=["SKIP"], inplace=True, errors="ignore")

    # --- If file doesn’t exist yet ---
    if not Path(review_path).exists():
        review_df.to_excel(review_path, index=False)
        return print(f"✅ Created new review log: {review_path}")

    # --- Open workbook and worksheet ---
    wb = load_workbook(review_path)
    ws = wb.active

    # --- Append new rows (retaining formatting) ---
    for r_idx, row in enumerate(dataframe_to_rows(review_df, index=False, header=False), start=ws.max_row + 1):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # --- Optional: Auto-fit column widths ---
    for col in ws.columns:
        max_length = max((len(str(cell.value)) if cell.value else 0) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    wb.save(review_path)
    print(f"✅ Processed items appended (formatting retained): {review_path}")



def update_icsp(icsp_df, lookup_df):
    """Apply replacements for ICSP file."""
    for _, row in lookup_df.iterrows():
        old, new = str(row["old"]).strip(), str(row["new"]).strip()

        if new == "NLA":
            icsp_df.loc[icsp_df["prod"] == old, "descrip2"] = "NO LONGER AVAILABLE"
        else:
            icsp_df.loc[icsp_df["prod"] == old, "descrip2"] = f"REPLD BY {new}"
            icsp_df.loc[icsp_df["prod"] == new, "descrip2"] = f"REPLS {old}"
    return icsp_df


def fetch_prodline_map():
    """Load old_prodline -> lk_prodline map from exported CSV instead of SQL."""
    csv_path = PATHS["sups_nlas_data"] / "Prodline_xref.csv"
    df = pd.read_csv(csv_path)
    return df

def update_icsw(icsw_df, lookup_df, prodline_map):
    """Apply replacement logic for ICSW file with safe fallbacks, including NLA handling."""

    for _, row in lookup_df.iterrows():
        old, new = str(row["old"]).strip(), str(row["new"]).strip()

        # --- Old numbers ---
        icsw_df.loc[icsw_df["prod"] == old, "vendprod"] = "!999!"

        # find the old part’s prodline
        if (icsw_df["prod"] == old).any():
            old_line = icsw_df.loc[icsw_df["prod"] == old, "prodline"].values[0]

            # if that old_line exists in prodline_map, swap it
            matches = prodline_map.loc[prodline_map["all_prodline"] == old_line, "lk_prodline"]
            if not matches.empty:
                new_line = matches.iloc[0]
                icsw_df.loc[icsw_df["prod"] == old, "prodline"] = new_line
            # else leave prodline unchanged

        # --- NLA case ---
        if new == "NLA":
            # Do NOT touch prodline, just ensure vendprod is marked
            continue  

        # --- New numbers ---
        # if vendprod = !999! → clear
        mask = (icsw_df["prod"] == new) & (icsw_df["vendprod"] == "!999!")
        icsw_df.loc[mask, "vendprod"] = ""

        # if prodline like %LK → swap for most_popular_prodline
        mask = (icsw_df["prod"] == new) & (icsw_df["prodline"].str.endswith("LK"))
        if mask.any():
            for op in icsw_df.loc[mask, "prodline"].unique():
                matches = prodline_map.loc[prodline_map["all_prodline"] == op, "most_popular_prodline"]
                if not matches.empty:
                    new_val = matches.iloc[0]
                    icsw_df.loc[(icsw_df["prod"] == new) & (icsw_df["prodline"] == op), "prodline"] = new_val
                # else leave prodline unchanged

        # old prodline in exceptions → force back
        old_prodline_vals = icsw_df.loc[icsw_df["prod"] == old, "prodline"].unique()
        if any(op in ["WCMSDC", "WPMSBR", "WPMSDC"] for op in old_prodline_vals):
            icsw_df.loc[icsw_df["prod"] == new, "prodline"] = old_prodline_vals[0]

        # --- New numbers ---
        if new != "NLA":
            # if vendprod = !999! → clear
            mask = (icsw_df["prod"] == new) & (icsw_df["vendprod"] == "!999!")
            icsw_df.loc[mask, "vendprod"] = ""

            # if prodline like %LK → swap for most_popular_prodline
            mask = (icsw_df["prod"] == new) & (icsw_df["prodline"].str.endswith("LK"))
            if mask.any():
                for op in icsw_df.loc[mask, "prodline"].unique():
                    matches = prodline_map.loc[prodline_map["all_prodline"] == op, "most_popular_prodline"]
                    if not matches.empty:
                        new_val = matches.iloc[0]  # take first match only
                        icsw_df.loc[(icsw_df["prod"] == new) & (icsw_df["prodline"] == op), "prodline"] = new_val

            # old prodline in exceptions → force back
            if icsw_df.loc[icsw_df["prod"] == old, "prodline"].isin(["WCMSDC", "WPMSBR", "WPMSDC"]).any():
                icsw_df.loc[icsw_df["prod"] == new, "prodline"] = \
                    icsw_df.loc[icsw_df["prod"] == old, "prodline"].values[0]

    return icsw_df


def check_additional_whse(icsw_df, lookup_df):
    """Return list of new#s needing additional warehouse rows (<29)."""
    report = []

    for _, row in lookup_df.iterrows():
        new = row["new"]
        if new != "NLA":
            count = (icsw_df["prod"] == new).sum()
            if count < 29:
                report.append(new)

    if report:
        print("Additional Warehouses needed for:", ", ".join(report))

    return report


def check_old_no_match(icsw_df, lookup_df):
    """Return list of reported old #s that did not appear in ICSW SAAMM."""
    report = []

    for _, row in lookup_df.iterrows():
        old = row["old"]
        count = (icsw_df["prod"] == old).sum()
        if count < 1:
            report.append(old)

    if report:
        print("Old products not found in ICSW:", ", ".join(report))

    return report

def autofit_worksheet(ws, padding=2):
    """Auto-fit column widths based on max cell length."""
    for col_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col_cells[0].column)

        for cell in col_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        ws.column_dimensions[col_letter].width = max_length + padding

def core_vend_nla_check(icsw_df, lookup_df):
    CORE_VEND = {"360", "775", "797", "825", "190", "298", "680", "784"}
    REPORT_PATH = PATHS["sups_nlas_data"] / "OEM NLAs Report.xlsx"
    SQLITE_PATH = PATHS["purchdata"]

    # --- 1. Identify NLA products from lookup ---
    nla_prods = (
        lookup_df.loc[lookup_df["new"].str.upper() == "NLA", "old"]
        .astype(str)
        .str.strip()
        .unique()
    )

    if len(nla_prods) == 0:
        print("✅ No NLA items in lookup.")
        return

    # --- 2. Filter ICSW for core vendors ---
    nla_icsw = icsw_df[
        (icsw_df["prod"].isin(nla_prods)) &
        (icsw_df["arpvendno"].isin(CORE_VEND))
    ][["prod", "arpvendno"]].drop_duplicates()

    if nla_icsw.empty:
        print("✅ No core NLAs found.")
        return

    print(f"⚠️ Found {len(nla_icsw)} core vendor NLA item(s). Querying SQLite to check usage...")

    # --- 3. Query SQLite ---
    placeholders = ",".join("?" for _ in nla_icsw["prod"].unique())
    sql = f"""
            select 
            DATE('now') AS DATE,
            arpvendno AS VENDOR,
            prod AS PRODUCT,
            vendprod AS "VENDOR PRODUCT",
            'NO LONER AVAILABLE' AS CHANGE,
            Total_Usage_12_Mo AS "TOTAL 12 MO USAGE 25",
            '' AS NOTES
            from csdusage
            WHERE prod in ({placeholders})
            --WHERE prod in ('771-6','WB31K10265','W10193094')
            AND mths_w_usge_count >= 4
            AND Total_Usage_12_Mo >= 4
            AND WAREHOUSE = '25'
            AND arpvendno in (360,775,797,825,190,298,680,784)
    """

    with sqlite3.connect(SQLITE_PATH) as conn:
        result_df = pd.read_sql_query(
            sql,
            conn,
            params=list(nla_icsw["prod"].unique())
        )

    if result_df.empty:
        print("ℹ️ SQLite query returned no rows.")
        return

  # --- 4. Append to Excel report ---
    if not REPORT_PATH.exists():
        result_df.to_excel(REPORT_PATH, index=False)

        wb = load_workbook(REPORT_PATH)
        ws = wb.active
        autofit_worksheet(ws)
        wb.save(REPORT_PATH)

        print(f"📄 Created OEM NLA report: {REPORT_PATH}")
        return


    wb = load_workbook(REPORT_PATH)
    ws = wb.active

    for r_idx, row in enumerate(
        dataframe_to_rows(result_df, index=False, header=False),
        start=ws.max_row + 1
    ):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    autofit_worksheet(ws)
    wb.save(REPORT_PATH)

    print(f"📄 Appended {len(result_df)} row(s) to OEM NLA report.")


# === MAIN ===
def main():

    # Get newest files
    icsp_file = get_newest_file("mmicsp")
    icsw_file = get_newest_file("mmicsw")
    lookup_file = PATHS["sups_nlas_data"] / "NLAs and SUPs to update via SAAMM.xlsx"

    # Load
    icsp_df = pd.read_csv(icsp_file, sep="\t", dtype=str)
    icsw_df = pd.read_csv(icsw_file, sep="\t", dtype=str)

    # strip spaces in all columns
    icsp_df = icsp_df.apply(lambda x: x.str.strip() if x.dtype == "object" else x)
    icsw_df = icsw_df.apply(lambda x: x.str.strip() if x.dtype == "object" else x)

    # 🔑 Load lookup file here
    lookup_df = load_lookup(lookup_file)

    #--- Add OEM NLAs to dedicated review file
    core_vend_nla_check(icsw_df, lookup_df)

    # Update ICSP
    icsp_df = update_icsp(icsp_df, lookup_df)

    # Update ICSW
    prodline_map = fetch_prodline_map()
    icsw_df = update_icsw(icsw_df, lookup_df, prodline_map)

    # Validation
    additional_whse_needed = check_additional_whse(icsw_df, lookup_df)
    old_no_match = check_old_no_match(icsw_df, lookup_df)
    

    # Save back out
    out = PATHS["sups_nlas_data"]

    icsp_out = os.path.join(out, "mmicsp_updated.txt")
    icsw_out = os.path.join(out, "mmicsw_updated.txt")

    icsp_df.to_csv(icsp_out, sep="\t", index=False)
    icsw_df.to_csv(icsw_out, sep="\t", index=False)

        # --- Append processed items to supervisor review file ---

    review_file = PATHS["sups_nlas_data"] / "NLAs and SUPs to process.xlsx"
    append_to_review_log(lookup_df,
     review_file)
    

    print("Files updated in SUP & NLA subfolder of Analyst Files!  Review file updated!")

        # --- Email notifications ---

    # --- Email: Issue Alerts ---
    if additional_whse_needed or old_no_match:

        subject = "⚠️ Supersede Process Alerts"

        issues = []

        if old_no_match:
            issues.append(
                "Old products not found in ICSW: " +
                ", ".join(old_no_match)
            )

        if additional_whse_needed:
            issues.append(
                "Additional warehouses needed for new products: " +
                ", ".join(additional_whse_needed)
            )

        body = (
            "Hello,\n\n"
            + "\n\n".join(issues)
            + "\n\nPlease review and take appropriate action.\n\n"
            "-- Automated SUP/NLA Process"
        )

        send_email(
            subject=subject,
            body=body,
            to_addrs=EMAILS["mass_maint_user"],
            cc=[],
            bcc=[],
            html=False
        )

    subject = "SUP/NLA Process List Updated Notification"
    body = (
        "Hello,\n\n"
        "This is an automated email to notify you that the 'NLAs and SUPs to process.xlsx' list has been updated.\n"
    )
    to_addrs = EMAILS["sup_nla_notification_emails"]
    cc = EMAILS["mass_maint_user"]
    bcc = []  # optional

    send_email(
        subject=subject,
        body=body,
        to_addrs=to_addrs,
        cc=cc,
        bcc=bcc,
        html=True  # nice formatting
    )
    print("Email notification sent!")
    print("✅ Process complete!")


if __name__ == "__main__":
    main()